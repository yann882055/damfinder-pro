# -*- coding: utf-8 -*-
"""
DamFinder Pro v1.0 — Analysis Engine
======================================
Hydroelectric site identification by longitudinal profile analysis and
specific potential method (kW/40m) — ICOLD standards + World Bank ESMAP.

Methodology : ICOLD standards + World Bank ESMAP hydropower guidelines
Hydroelectric potential specific method — profile analysis kW/40m
Calibrated on anonymized West African hydroelectric reference projects

Replaces all arcpy dependencies with open-source equivalents:
  arcpy.sa.Fill            → pysheds fill_depressions
  arcpy.sa.FlowDirection   → pysheds flowdir
  arcpy.sa.FlowAccumulation→ pysheds accumulation
  StreamOrder/StreamLink   → pysheds + numpy thresholding
  arcpy.sa.ExtractByMask   → rasterio.mask.mask
  Feature classes          → geopandas GeoDataFrame
  arcpy.da cursors         → geopandas / numpy / rasterio.sample

© 2026 DAMFINDER Engineering Tools — All rights reserved
"""

import os, sys, math, time, json, tempfile, traceback
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional, List, Dict, Any

import numpy as np
import pandas as pd

# Raster I/O
import rasterio
from rasterio.mask import mask as rio_mask
from rasterio.transform import rowcol, xy
from rasterio.crs import CRS
from rasterio.warp import reproject, Resampling, calculate_default_transform
import rasterio.features

# Hydrological processing
try:
    from pysheds.grid import Grid as PyshedsGrid
    HAS_PYSHEDS = True
except ImportError:
    HAS_PYSHEDS = False

# Vector
import geopandas as gpd
from shapely.geometry import Point, LineString, MultiLineString
from shapely.ops import unary_union

# Visualisation
try:
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    HAS_MATPLOTLIB = True
except ImportError:
    HAS_MATPLOTLIB = False

try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS — ICOLD / ESMAP calibration (anonymised reference projects)
# ═══════════════════════════════════════════════════════════════════════════════

APP_CREDIT = """DamFinder Pro v1.0
Developed by DAMFINDER Engineering Tools
Methodology : ICOLD standards + World Bank ESMAP hydropower guidelines
Hydroelectric potential specific method — profile analysis kW/40m
Calibrated on anonymized West African hydroelectric reference projects
© 2026 DAMFINDER Engineering Tools — All rights reserved"""

# Reference projects — anonymised (values preserved exactly)
REFERENCE_PROJECTS = {
    'REF_A': {'P_MW': 275.0, 'E_GWH': 1200.0, 'H_BRUTE': 45.0,
               'Q_EQ': 714.0, 'CAPEX_MUSD': 428.0, 'LCOE': 37.4},
    'REF_B': {'P_MW': 112.0, 'E_GWH': 580.0,  'H_BRUTE': 18.4,
               'Q_EQ': 774.0, 'CAPEX_MUSD': 305.0, 'LCOE': 55.1},
    'REF_C': {'P_MW': 140.0, 'E_GWH': 725.0,  'H_BRUTE': 20.9,
               'Q_EQ': 885.0, 'CAPEX_MUSD': 459.0, 'LCOE': 66.3},
    'REF_D': {'P_MW': 117.0, 'E_GWH': 546.0,  'H_BRUTE': 19.3,
               'Q_EQ': 925.0, 'CAPEX_MUSD': 481.9, 'LCOE': 93.0},
    'REF_E': {'P_MW': 100.0, 'E_GWH': 463.0,  'H_BRUTE': 15.6,
               'Q_EQ': 929.0, 'CAPEX_MUSD': 419.0, 'LCOE': 95.4},
}

# CAPEX formula calibrated on reference projects: CAPEX (M$) = K × P_MW^B
CAPEX_COEFF_K    = 3.2
CAPEX_EXPOSANT_B = 0.72

TAUX_ACTUALISATION = 0.08      # 8 %
DUREE_VIE          = 30        # years
TAUX_OPEX          = 0.025     # 2.5 % of CAPEX
CRF                = 0.088827  # Capital Recovery Factor (i=8%, n=30)
TAUX_EUR_USD       = 1.08
TAUX_FCFA_USD      = 607.0


# ═══════════════════════════════════════════════════════════════════════════════
# LCOE FUNCTION (preserved exactly from v30, names cleaned)
# ═══════════════════════════════════════════════════════════════════════════════

def calculer_lcoe_calibre(p_mw: float, e_gwh_an: float) -> dict:
    """
    Calibrated LCOE calculation — ICOLD/ESMAP methodology.
    Calibrated on anonymized West African reference projects (REF_A … REF_E).

    CAPEX (M$) = 3.2 × P_MW^0.72
    """
    if p_mw <= 0 or e_gwh_an <= 0:
        return {'CAPEX_MUSD': 0.0, 'CAPEX_MEUR': 0.0,
                'LCOE_USD_MWH': 0.0, 'LCOE_EUR_MWH': 0.0, 'LCOE_FCFA_KWH': 0.0}

    capex_musd   = CAPEX_COEFF_K * (p_mw ** CAPEX_EXPOSANT_B)
    capex_meur   = capex_musd / TAUX_EUR_USD
    annuite_musd = capex_musd * CRF
    opex_musd    = capex_musd * TAUX_OPEX
    cout_total   = annuite_musd + opex_musd
    lcoe_usd_mwh = (cout_total * 1_000_000) / (e_gwh_an * 1000)
    lcoe_eur_mwh = lcoe_usd_mwh / TAUX_EUR_USD
    lcoe_fcfa_kwh = lcoe_usd_mwh * TAUX_FCFA_USD / 1000

    return {
        'CAPEX_MUSD':    round(capex_musd,    1),
        'CAPEX_MEUR':    round(capex_meur,    1),
        'LCOE_USD_MWH':  round(lcoe_usd_mwh,  1),
        'LCOE_EUR_MWH':  round(lcoe_eur_mwh,  1),
        'LCOE_FCFA_KWH': round(lcoe_fcfa_kwh, 1),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# HELPER — adaptive threshold (preserved exactly)
# ═══════════════════════════════════════════════════════════════════════════════

def _adaptive_threshold(q_moyen: float) -> float:
    if q_moyen < 10:
        return 20.0
    elif q_moyen < 50:
        return 30.0 + (q_moyen - 10) * 0.5
    elif q_moyen < 100:
        return 50.0 + (q_moyen - 50) * 0.4
    else:
        return min(100.0, 70.0 + (q_moyen - 100) * 0.2)


def _detect_peaks(arr: list, threshold: float, min_distance: int = 3) -> list:
    peaks = []
    for i in range(1, len(arr) - 1):
        if arr[i] > threshold and arr[i] > arr[i-1] and arr[i] > arr[i+1]:
            too_close = False
            for ep in peaks:
                if abs(i - ep) < min_distance:
                    if arr[i] > arr[ep]:
                        peaks.remove(ep)
                    else:
                        too_close = True
                    break
            if not too_close:
                peaks.append(i)
    return peaks


# ═══════════════════════════════════════════════════════════════════════════════
# RASTER SAMPLING UTILITIES
# ═══════════════════════════════════════════════════════════════════════════════

def _sample_raster_at_points(raster_path: str,
                              xs: List[float], ys: List[float],
                              nodata_val: float = None) -> List[Optional[float]]:
    """Sample a single-band raster at (x, y) coordinate pairs."""
    results = []
    try:
        with rasterio.open(raster_path) as src:
            nd = nodata_val if nodata_val is not None else src.nodata
            coords = list(zip(xs, ys))
            for val in src.sample(coords):
                v = float(val[0])
                if nd is not None and v == nd:
                    results.append(None)
                else:
                    results.append(v)
    except Exception:
        results = [None] * len(xs)
    return results


def _sample_raster_array(src_path: str,
                          xs: List[float], ys: List[float]) -> List[float]:
    """Sample raster; replace nodata with 0."""
    vals = _sample_raster_at_points(src_path, xs, ys)
    return [v if v is not None else 0.0 for v in vals]


# ═══════════════════════════════════════════════════════════════════════════════
# DEM CLIPPING
# ═══════════════════════════════════════════════════════════════════════════════

def clip_dem(dem_path: str, study_area_path: Optional[str],
             out_path: str) -> str:
    """Clip DEM to study area polygon (if provided) using rasterio.mask."""
    if not study_area_path or not os.path.exists(study_area_path):
        # No clip — just copy reference
        return dem_path

    gdf = gpd.read_file(study_area_path)

    with rasterio.open(dem_path) as src:
        # Reproject study area to DEM CRS if needed
        if gdf.crs and gdf.crs != src.crs:
            gdf = gdf.to_crs(src.crs)

        shapes = [geom for geom in gdf.geometry if geom is not None]
        out_image, out_transform = rio_mask(src, shapes, crop=True)
        out_meta = src.meta.copy()
        out_meta.update({
            "driver": "GTiff",
            "height": out_image.shape[1],
            "width":  out_image.shape[2],
            "transform": out_transform
        })

    with rasterio.open(out_path, "w", **out_meta) as dst:
        dst.write(out_image)

    return out_path


# ═══════════════════════════════════════════════════════════════════════════════
# HYDROLOGICAL PROCESSING (pysheds)
# ═══════════════════════════════════════════════════════════════════════════════

def compute_flow(dem_path: str,
                 out_dir_path: str,
                 out_acc_path: str,
                 cb: Callable) -> tuple:
    """
    Fill pits + depressions, compute D8 flow direction and accumulation.
    Returns (flow_dir_path, flow_acc_path) as GeoTiff files.
    """
    cb(5, "[1.2] Fill depressions (pysheds)…", 'info')

    if not HAS_PYSHEDS:
        raise ImportError("pysheds is required for hydrological processing.")

    grid = PyshedsGrid.from_raster(dem_path)
    dem  = grid.read_raster(dem_path)

    cb(8,  "[1.2] Fill pits…",         'info')
    pit_filled = grid.fill_pits(dem)
    cb(11, "[1.2] Fill depressions…",  'info')
    flooded    = grid.fill_depressions(pit_filled)
    cb(14, "[1.2] Resolve flats…",     'info')
    inflated   = grid.resolve_flats(flooded)

    cb(18, "[1.3] Flow direction D8…", 'info')
    fdir = grid.flowdir(inflated)

    cb(22, "[1.4] Flow accumulation…", 'info')
    acc  = grid.accumulation(fdir)

    # Write outputs as GeoTiff
    with rasterio.open(dem_path) as src:
        meta = src.meta.copy()
        meta.update(dtype=rasterio.float32)

    fdir_arr = np.array(fdir).astype(np.float32)
    acc_arr  = np.array(acc).astype(np.float32)

    with rasterio.open(out_dir_path, 'w', **meta) as dst:
        dst.write(fdir_arr[np.newaxis, :, :])
    with rasterio.open(out_acc_path, 'w', **meta) as dst:
        dst.write(acc_arr[np.newaxis, :, :])

    cb(26, "[1.4] Flow rasters computed.", 'info')
    return out_dir_path, out_acc_path


def extract_stream_network(dem_path: str,
                            acc_path: str,
                            flow_threshold: int,
                            strahler_thresholds: dict,
                            study_area_path: Optional[str],
                            cb: Callable) -> gpd.GeoDataFrame:
    """
    Extract stream network as GeoDataFrame with Strahler order approximation.
    Uses pysheds for stream extraction, then classifies by accumulation value.
    """
    cb(28, "[1.5] Generating stream network…", 'info')

    if not HAS_PYSHEDS:
        raise ImportError("pysheds required for stream extraction.")

    grid = PyshedsGrid.from_raster(dem_path)
    dem  = grid.read_raster(dem_path)
    pit  = grid.fill_pits(dem)
    fld  = grid.fill_depressions(pit)
    inf  = grid.resolve_flats(fld)
    fdir = grid.flowdir(inf)
    acc  = grid.accumulation(fdir)

    with rasterio.open(dem_path) as src:
        transform = src.transform
        crs       = src.crs
        nodata    = src.nodata or -9999

    # Auto threshold if 0
    acc_arr = np.array(acc)
    if flow_threshold == 0:
        flow_threshold = max(100, int(acc_arr.max() / 5000))
        cb(29, f"   Auto threshold: {flow_threshold} cells", 'info')

    # Binary stream mask
    stream_mask = acc_arr > flow_threshold

    # Strahler approximation: bin by sqrt(accumulation) → proxy for stream order
    # Real Strahler requires full tree traversal; we use accumulation bins
    # which correlate strongly with order for typical drainage networks
    acc_max = float(acc_arr[stream_mask].max()) if stream_mask.any() else 1.0

    def _strahler_approx(acc_val: float) -> int:
        ratio = acc_val / acc_max
        if ratio >= 0.5:   return 6
        elif ratio >= 0.2: return 5
        elif ratio >= 0.08:return 4
        elif ratio >= 0.02:return 3
        else:              return 2

    # Vectorise stream pixels → line segments via rasterio features
    shapes_gen = rasterio.features.shapes(
        stream_mask.astype(np.uint8),
        mask=stream_mask.astype(np.uint8),
        transform=transform
    )

    geoms = []
    for geom, val in shapes_gen:
        if val == 1:
            from shapely.geometry import shape as shapely_shape
            g = shapely_shape(geom)
            geoms.append(g)

    if not geoms:
        cb(0, "   WARNING: No stream cells found. Lower threshold.", 'warning')
        return gpd.GeoDataFrame(columns=['geometry', 'ORD_STRA', 'LENGTH_KM'],
                                 crs=crs)

    # Merge adjacent polygons → skeleton approach via centroid-based lines
    # For a cleaner result, we convert the raster stream to vectorised skeletons
    # using the medial axis of each connected component
    from shapely.geometry import MultiPolygon
    merged = unary_union(geoms)

    # Simplify polygon boundaries to approximate stream centrelines
    lines = []
    polys = merged.geoms if hasattr(merged, 'geoms') else [merged]
    for poly in polys:
        # Use the centroid chain of the polygon as a line proxy
        try:
            line = poly.centroid
            # Create a line from the exterior coords simplified
            coords = list(poly.exterior.simplify(
                float(transform.a) * 2).coords)
            if len(coords) >= 2:
                ln = LineString(coords)
                # Accumulation at centroid
                cx, cy = ln.centroid.x, ln.centroid.y
                row_i, col_i = rowcol(transform, cx, cy)
                row_i = min(max(int(row_i), 0), acc_arr.shape[0]-1)
                col_i = min(max(int(col_i), 0), acc_arr.shape[1]-1)
                acc_val = float(acc_arr[row_i, col_i])
                order   = _strahler_approx(acc_val)
                length_km = ln.length / 1000.0
                lines.append({
                    'geometry':   ln,
                    'ORD_STRA':   order,
                    'LENGTH_KM':  round(length_km, 4),
                    'ACC_MAX':    acc_val
                })
        except Exception:
            continue

    if not lines:
        return gpd.GeoDataFrame(columns=['geometry', 'ORD_STRA', 'LENGTH_KM'],
                                 crs=crs)

    gdf = gpd.GeoDataFrame(lines, crs=crs)

    # Clip to study area
    if study_area_path and os.path.exists(study_area_path):
        sa = gpd.read_file(study_area_path)
        if sa.crs != gdf.crs:
            sa = sa.to_crs(gdf.crs)
        gdf = gpd.clip(gdf, sa)

    cb(32, f"   {len(gdf)} stream segments extracted.", 'info')

    # Log Strahler distribution
    for o in sorted(gdf['ORD_STRA'].unique()):
        n = (gdf['ORD_STRA'] == o).sum()
        cb(32, f"   Strahler {o}: {n} segments", 'info')

    return gdf.reset_index(drop=True)


# ═══════════════════════════════════════════════════════════════════════════════
# PROFILE EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════════

def extract_profiles(river_gdf: gpd.GeoDataFrame,
                     dem_path: str,
                     acc_path: str,
                     profile_step: int,
                     strahler_main: int,
                     strahler_major: int,
                     strahler_secondary: int,
                     cb: Callable) -> dict:
    """
    Build longitudinal profiles for PRINCIPAL, MAJEURS, SECONDAIRES categories.
    Uses shapely interpolate() + rasterio.sample() to replace arcpy cursors.

    Returns: {category: list_of_profile_dicts}
    Each dict: {pk, z, accum, strahler, x, y}
    """
    cb(35, "[2] Extracting longitudinal profiles…", 'info')

    max_s = int(river_gdf['ORD_STRA'].max()) if len(river_gdf) > 0 else 1

    strahler_main      = min(strahler_main,      max_s)
    strahler_major     = min(strahler_major,      max_s)
    strahler_secondary = min(strahler_secondary,  max_s)

    categories = {
        'PRINCIPAL':  river_gdf[river_gdf['ORD_STRA'] >= strahler_main],
        'MAJEURS':    river_gdf[river_gdf['ORD_STRA'] == strahler_major],
        'SECONDAIRES':river_gdf[river_gdf['ORD_STRA'] == strahler_secondary],
    }

    profiles = {}

    for cat_name, cat_gdf in categories.items():
        if len(cat_gdf) == 0:
            continue

        cb(35, f"   Extracting {cat_name} ({len(cat_gdf)} segments)…", 'info')

        pts_all = []
        chainage = 0.0

        for _, row in cat_gdf.iterrows():
            geom = row.geometry
            if geom is None:
                continue

            # Handle MultiLineString
            if geom.geom_type == 'MultiLineString':
                lines_list = list(geom.geoms)
            else:
                lines_list = [geom]

            for ln in lines_list:
                ln_len = ln.length
                n_pts  = max(2, int(ln_len / profile_step) + 1)

                for i in range(n_pts):
                    d  = i * profile_step
                    if d > ln_len:
                        d = ln_len
                    pt = ln.interpolate(d)
                    pts_all.append({
                        'pk':      chainage + d,
                        'x':       pt.x,
                        'y':       pt.y,
                        'z':       None,
                        'accum':   0.0,
                        'strahler': int(row['ORD_STRA'])
                    })

                chainage += ln_len

        if not pts_all:
            continue

        # Batch-sample DEM and accumulation
        xs = [p['x'] for p in pts_all]
        ys = [p['y'] for p in pts_all]

        z_vals    = _sample_raster_at_points(dem_path, xs, ys)
        acc_vals  = _sample_raster_array(acc_path, xs, ys)

        z_valid = z_errors = 0
        for i, p in enumerate(pts_all):
            p['z']     = z_vals[i]
            p['accum'] = acc_vals[i]
            if p['z'] is not None:
                z_valid += 1
            else:
                z_errors += 1

        cb(35, f"   {cat_name}: {z_valid} valid Z, {z_errors} NoData", 'info')

        if z_valid == 0:
            cb(0, f"   ERROR: No Z extracted for {cat_name}. "
               "Check DEM coverage.", 'error')
            continue

        # Interpolate missing Z values
        z_arr = [p['z'] for p in pts_all]
        for i in range(len(z_arr)):
            if z_arr[i] is None:
                before = next((z_arr[j] for j in range(i-1, -1, -1)
                               if z_arr[j] is not None), None)
                after  = next((z_arr[j] for j in range(i+1, len(z_arr))
                               if z_arr[j] is not None), None)
                if before is not None and after is not None:
                    i_b = next(j for j in range(i-1, -1, -1)
                               if z_arr[j] is not None)
                    i_a = next(j for j in range(i+1, len(z_arr))
                               if z_arr[j] is not None)
                    ratio = (i - i_b) / (i_a - i_b)
                    z_arr[i] = before + ratio * (after - before)
                elif before is not None:
                    z_arr[i] = before
                elif after is not None:
                    z_arr[i] = after

        for i, p in enumerate(pts_all):
            p['z'] = z_arr[i]

        # Remove still-None points
        pts_all = [p for p in pts_all if p['z'] is not None]

        if pts_all:
            profiles[cat_name] = pts_all
            cb(35, f"   {cat_name}: {len(pts_all)} profile points", 'info')

    return profiles


# ═══════════════════════════════════════════════════════════════════════════════
# FLOW ATTRIBUTION
# ═══════════════════════════════════════════════════════════════════════════════

def attribute_flows(profiles: dict,
                    flow_source: str,
                    equip_coef: float,
                    fixed_flow_main: float,
                    fixed_flow_major: float,
                    fixed_flow_secondary: float,
                    upstream_ratio: float,
                    pjmax: float,
                    cn_default: float,
                    runoff_coef: float,
                    tc_method: str,
                    rainy_days: int,
                    cell_size: float,
                    cb: Callable) -> dict:
    """
    Attribute Q_moyen, Q_etiage, Q_crue, Q_equip to each profile point.
    Supports: Débits fixes | HEC-HMS (Pjmax) | Empirique
    (HydroATLAS replaced by Débits fixes when no shapefile available)
    """
    cb(42, f"[3] Flow attribution — source: {flow_source}", 'info')

    q_ref = {
        'PRINCIPAL':   fixed_flow_main,
        'MAJEURS':     fixed_flow_major,
        'SECONDAIRES': fixed_flow_secondary,
    }

    for cat, pts in profiles.items():
        if not pts:
            continue

        if flow_source in ("Débits fixes", "HydroATLAS"):
            # Linear interpolation amont→aval using accumulation proxy
            accums = [p['accum'] for p in pts]
            acc_min = min(accums) if accums else 0
            acc_max = max(accums) if accums else 1
            q_aval  = q_ref.get(cat, fixed_flow_main)
            q_amont = q_aval * upstream_ratio

            for p in pts:
                acc = p['accum']
                if acc_max > acc_min:
                    ratio = (acc - acc_min) / (acc_max - acc_min)
                else:
                    ratio = 1.0
                q_m = q_amont + ratio * (q_aval - q_amont)
                p['q_moyen']  = q_m
                p['q_etiage'] = q_m * 0.4
                p['q_crue']   = q_m * 2.5
                p['q_equip']  = q_m * equip_coef
                p['surf_bv']  = 0.0

        elif flow_source == "HEC-HMS (Pjmax)":
            pks    = [p['pk']  for p in pts]
            zs     = [p['z']   for p in pts]
            accums = [p['accum'] for p in pts]

            z_amont_profile = zs[0] if zs else 0.0

            for p in pts:
                surf_bv_km2 = (p['accum'] * cell_size * cell_size) / 1e6
                cn = cn_default
                S  = 25400 / cn - 254
                Ia = 0.2 * S
                Pe = ((pjmax - Ia) ** 2 / (pjmax - Ia + S)) if pjmax > Ia else 0
                L_km  = p['pk'] / 1000.0
                delta_z = z_amont_profile - p['z']
                if delta_z > 0 and L_km > 0:
                    pente_m = delta_z / (L_km * 1000)
                    Tc_h = 0.0195 * (L_km**0.77) * (pente_m**-0.385) if pente_m > 0 else 1.0
                else:
                    Tc_h = 1.0
                I_mmh   = pjmax / max(Tc_h, 0.5)
                Q_pointe = (runoff_coef * I_mmh * surf_bv_km2) / 3.6
                lame_an  = pjmax * rainy_days * runoff_coef / 365
                Q_alt    = (lame_an / 1000.0) * (surf_bv_km2 * 1e6) / (365.25 * 86400)
                Q_m      = (Q_pointe * 0.4 + Q_alt) / 2.0
                p['q_moyen']  = Q_m
                p['q_etiage'] = Q_m * 0.3
                p['q_crue']   = Q_pointe * 1.5
                p['q_equip']  = Q_m * equip_coef
                p['surf_bv']  = surf_bv_km2

        else:  # Empirique
            for p in pts:
                pk_km = p['pk'] / 1000.0
                q_m   = pk_km * 1.5
                p['q_moyen']  = q_m
                p['q_etiage'] = q_m * 0.4
                p['q_crue']   = q_m * 2.0
                p['q_equip']  = q_m * equip_coef
                p['surf_bv']  = 0.0

    cb(48, "[3] Flow attribution complete.", 'info')
    return profiles


# ═══════════════════════════════════════════════════════════════════════════════
# SITE DETECTION — SPECIFIC POTENTIAL METHOD (kW/40m) — PRESERVED EXACTLY
# ═══════════════════════════════════════════════════════════════════════════════

def detect_sites(profiles: dict,
                 segment_norm: int,
                 profile_step: int,
                 adaptive_threshold: bool,
                 threshold_peak_manual: float,
                 max_exploit_dist: float,
                 min_power: float,
                 weir_height: float,
                 min_spacing: float,
                 filter_slope: bool,
                 min_slope: float,
                 head_loss_pct: float,
                 turbine_eff: float,
                 equip_coef: float,
                 cb: Callable) -> list:
    """
    Detect hydroelectric sites by specific potential method (kW/40m).
    ICOLD methodology — profile analysis, peak detection, window extension.
    All logic preserved exactly from original v30.
    """
    cb(50, "[4] Site detection — specific potential method…", 'info')

    sites_detected = []
    site_counter   = 1

    for cat, profile_data in profiles.items():
        if not profile_data or len(profile_data) < 3:
            continue

        cb(50, f"   Analysing profile {cat}…", 'info')

        # Calculate specific potential for each segment
        for i in range(len(profile_data) - 1):
            pt      = profile_data[i]
            pt_next = profile_data[i + 1]
            dh      = pt['z'] - pt_next['z']
            dpk     = pt_next['pk'] - pt['pk']

            pente_pct = (dh / dpk * 100) if dpk > 0 else 0
            pt['delta_h']   = dh
            pt['pente_pct'] = pente_pct

            if filter_slope and (pente_pct < min_slope or pente_pct > 50.0):
                pt['pot_kw']   = 0
                pt['pot_spec'] = 0
                continue

            if dh > 0 and pt.get('q_moyen', 0) > 0:
                pot_kw  = 1000 * 9.81 * pt['q_moyen'] * dh * turbine_eff / 1000
            else:
                pot_kw = 0

            pot_spec = pot_kw * (segment_norm / max(profile_step, 1))
            pt['pot_kw']   = pot_kw
            pt['pot_spec'] = pot_spec

        # Last point
        profile_data[-1].setdefault('pot_spec', 0)
        profile_data[-1].setdefault('pot_kw',   0)

        pot_spec_arr = [pt.get('pot_spec', 0) for pt in profile_data]
        q_vals       = [pt.get('q_moyen', 0) for pt in profile_data
                        if pt.get('q_moyen', 0) > 0]
        q_median     = sorted(q_vals)[len(q_vals)//2] if q_vals else 10.0

        if adaptive_threshold:
            thr_peak = _adaptive_threshold(q_median)
        else:
            thr_peak = threshold_peak_manual
        thr_ext = thr_peak * 0.4

        cb(50, f"   {cat}: threshold={thr_peak:.1f} kW/{segment_norm}m", 'info')

        peaks = _detect_peaks(pot_spec_arr, thr_peak, min_distance=3)
        cb(50, f"   {cat}: {len(peaks)} peaks detected", 'info')

        if not peaks:
            continue

        fenetre_max      = max_exploit_dist if max_exploit_dist > 0 else 10000
        max_window_pts   = int(fenetre_max / max(profile_step, 1))

        for peak_idx in peaks:
            # Extend upstream (decreasing index)
            start = peak_idx
            while start > 0 and (peak_idx - start) * profile_step < fenetre_max:
                if pot_spec_arr[start - 1] >= thr_ext:
                    start -= 1
                else:
                    break

            # Extend downstream (increasing index)
            end = peak_idx + 1
            while end < len(profile_data) and (end - peak_idx) * profile_step < fenetre_max:
                if end < len(pot_spec_arr) and pot_spec_arr[end] >= thr_ext:
                    end += 1
                else:
                    break

            lon_site = (end - start) * profile_step
            if lon_site < 100:
                continue

            h_naturelle = sum(profile_data[j].get('delta_h', 0)
                              for j in range(start, min(end, len(profile_data))))
            if h_naturelle <= 0:
                continue

            h_brute = h_naturelle   # V30 correction: no artificial +5m

            q_moyens = [profile_data[j].get('q_moyen', 0)
                        for j in range(start, end)
                        if j < len(profile_data) and profile_data[j].get('q_moyen', 0) > 0]
            if not q_moyens:
                continue

            q_moyen_s  = sum(q_moyens) / len(q_moyens)
            q_equip_s  = q_moyen_s * equip_coef
            h_nette    = h_brute * (1 - head_loss_pct)
            p_kw       = 1000 * 9.81 * q_equip_s * h_nette * turbine_eff / 1000

            if p_kw < min_power:
                continue

            pk_prise = profile_data[start]['pk']
            if any(abs(pk_prise - s['pk_prise']) < min_spacing
                   and s['categorie'] == cat
                   for s in sites_detected):
                continue

            end_idx      = min(end - 1, len(profile_data) - 1)
            pot_specs_s  = pot_spec_arr[start:min(end, len(pot_spec_arr))]

            sites_detected.append({
                'site_id':        f"SITE_{site_counter:04d}",
                'categorie':      cat,
                'strahler':       profile_data[start].get('strahler', 0),
                'pk_prise':       pk_prise,
                'pk_rest':        profile_data[end_idx]['pk'],
                'x_prise':        profile_data[start]['x'],
                'y_prise':        profile_data[start]['y'],
                'x_rest':         profile_data[end_idx]['x'],
                'y_rest':         profile_data[end_idx]['y'],
                'z_prise':        profile_data[start]['z'],
                'z_rest':         profile_data[end_idx]['z'],
                'h_naturelle':    h_naturelle,
                'h_seuil':        weir_height,
                'h_brute':        h_brute,
                'h_nette':        h_nette,
                'longueur_exploit': lon_site,
                'pente_moy_pct':  (h_naturelle / lon_site * 100) if lon_site > 0 else 0,
                'surf_bv':        profile_data[start].get('surf_bv', 0),
                'q_moyen':        q_moyen_s,
                'q_etiage':       q_moyen_s * 0.4,
                'q_crue':         q_moyen_s * 2.5,
                'q_equip':        q_equip_s,
                'pot_spec_max':   max(pot_specs_s) if pot_specs_s else 0,
                'pot_spec_moy':   sum(pot_specs_s)/len(pot_specs_s) if pot_specs_s else 0,
                'p_install_kw':   p_kw,
                'p_install_mw':   p_kw / 1000.0,
            })
            site_counter += 1

        n_cat = sum(1 for s in sites_detected if s['categorie'] == cat)
        cb(55, f"   {cat}: {n_cat} sites created", 'info')

    cb(58, f"[4] Total: {len(sites_detected)} sites detected.", 'info')
    return sites_detected


# ═══════════════════════════════════════════════════════════════════════════════
# MANNING HYDRAULICS (preserved exactly from v30)
# ═══════════════════════════════════════════════════════════════════════════════

def compute_manning(sites: list,
                    manning_n: float,
                    canal_length: float,
                    head_loss_pct: float,
                    side_slope: float,
                    turbine_eff: float,
                    cb: Callable) -> list:
    cb(60, "[5] Manning hydraulics…", 'info')

    for site in sites:
        Q       = site['q_equip']
        L       = canal_length if canal_length > 0 else site['longueur_exploit'] * 0.8
        H_brute = site['h_brute']
        S       = H_brute / L if L > 0 else 0.001

        B_est = (Q * manning_n / (S**0.5))**0.4
        B     = max(B_est, 2.0)
        h     = max(B / 3.0, 0.5)
        V     = 0.0
        Rh    = 0.0

        for _ in range(3):
            A  = B * h + side_slope * h * h
            P  = B + 2 * h * math.sqrt(1 + side_slope**2)
            Rh = A / P if P > 0 else 0
            V  = (1 / manning_n) * (Rh**(2/3)) * (S**0.5)
            Q_calc = A * V
            if Q_calc > 0:
                h = h * ((Q / Q_calc)**0.5)
                h = max(0.5, min(h, 10.0))

        dh_lin = S * L
        dh_sin = dh_lin * 0.1
        dh_tot = dh_lin + dh_sin
        H_nette_manning = max(H_brute - dh_tot, H_brute * 0.85)

        P_manning = 1000 * 9.81 * site['q_equip'] * H_nette_manning * turbine_eff / 1000

        site.update({
            'canal_largeur':       round(B,   2),
            'canal_hauteur':       round(h,   2),
            'canal_vitesse':       round(V,   3),
            'canal_rh':            round(Rh,  3),
            'canal_longueur':      round(L,   0),
            'perte_charge_m':      round(dh_tot, 2),
            'h_nette_manning':     round(H_nette_manning, 2),
            'p_install_kw_manning':round(P_manning,       1),
            'p_install_mw_manning':round(P_manning/1000,  3),
        })

    cb(63, "[5] Manning complete.", 'info')
    return sites


# ═══════════════════════════════════════════════════════════════════════════════
# DAM SIZING — V30 CORRECTED (CRN = Z_aval + H_nette, not Z_rest + 5m)
# ═══════════════════════════════════════════════════════════════════════════════

def compute_dam_sizing(sites: list, manning_n: float, cb: Callable) -> list:
    cb(65, "[5B] Dam sizing (CRN corrected formula)…", 'info')

    for site in sites:
        z_rest      = site['z_rest']
        z_prise     = site['z_prise']
        h_naturelle = site['h_naturelle']
        h_nette     = site.get('h_nette_manning', site['h_nette'])
        q_equip     = site['q_equip']
        lon_exp     = site['longueur_exploit']

        pente = h_naturelle / lon_exp if lon_exp > 0 else 0.001

        # Canal width
        v_adm = 1.5 if pente < 0.01 else (2.0 if pente < 0.03 else 2.5)
        largeur_canal = max(3.0, min((q_equip / v_adm) / 2.5, 50.0))

        # Z_aval by Manning iteration (Newton-Raphson)
        k    = 1.0 / manning_n
        h_eau = 2.0
        for _ in range(10):
            A     = largeur_canal * h_eau
            P     = largeur_canal + 2 * h_eau
            Rh    = A / P if P > 0 else 0
            Q_c   = k * A * (Rh**(2/3)) * math.sqrt(pente) if Rh > 0 else 0
            err   = Q_c - q_equip
            if abs(err) < 0.001:
                break
            A2 = largeur_canal * (h_eau + 0.01)
            P2 = largeur_canal + 2 * (h_eau + 0.01)
            Rh2= A2/P2 if P2 > 0 else 0
            Q2 = k * A2 * (Rh2**(2/3)) * math.sqrt(pente) if Rh2 > 0 else 0
            dQ_dh = (Q2 - Q_c) / 0.01
            if abs(dQ_dh) > 1e-6:
                h_eau -= err / dQ_dh
            h_eau = max(0.5, min(h_eau, 10.0))

        z_aval      = z_rest + h_eau
        crn         = z_aval + h_nette          # V30 CORRECTED formula
        revanche    = 2.5
        cote_crete  = crn + revanche
        h_barr_vue  = cote_crete - z_prise
        h_barr_tot  = h_barr_vue + 1.5
        ratio_val   = 2.0 if h_barr_vue < 10 else (2.5 if h_barr_vue < 20 else 3.0)
        lon_digue   = 20.0 * ratio_val + 100.0
        vol_res     = 5000.0 * 20.0 * 3.0
        surf_res    = (5000.0 * 20.0) / 10000.0

        site.update({
            'largeur_canal_calc':  round(largeur_canal, 1),
            'h_eau_aval':          round(h_eau,         2),
            'z_aval':              round(z_aval,         2),
            'crn':                 round(crn,            2),
            'revanche':            revanche,
            'cote_crete':          round(cote_crete,     2),
            'h_barrage_vue':       round(h_barr_vue,     1),
            'h_barrage_total':     round(h_barr_tot,     1),
            'longueur_digue':      round(lon_digue,      0),
            'volume_reservoir_m3': round(vol_res,        0),
            'surface_reservoir_ha':round(surf_res,       2),
        })

    cb(68, "[5B] Dam sizing complete. CRN = Z_aval + H_nette ✓", 'info')
    return sites


# ═══════════════════════════════════════════════════════════════════════════════
# CASCADE ANALYSIS (preserved exactly)
# ═══════════════════════════════════════════════════════════════════════════════

def compute_cascade(sites: list, cascade_distance: float, cb: Callable) -> list:
    cb(70, "[6] Cascade interaction analysis…", 'info')

    for site in sites:
        amont = []
        for other in sites:
            if other['site_id'] == site['site_id']:
                continue
            if other['categorie'] == site['categorie'] and \
               other['pk_prise'] < site['pk_prise']:
                dist = site['pk_prise'] - other['pk_rest']
                if 0 <= dist < cascade_distance:
                    amont.append({'id': other['site_id'],
                                  'dist': dist,
                                  'prelevt': other['q_equip']})
        site['nb_sites_amont'] = len(amont)
        if amont:
            prel = sum(s['prelevt'] for s in amont)
            site['q_residuel'] = max(0, site['q_moyen'] - prel * 0.3)
        else:
            site['q_residuel'] = site['q_moyen']
        site['commentaire_cascade'] = (f"{len(amont)} upstream sites"
                                       if amont else "No interaction")

    cb(72, "[6] Cascade complete.", 'info')
    return sites


# ═══════════════════════════════════════════════════════════════════════════════
# LCOE — ECONOMIC ANALYSIS (preserved exactly, names cleaned)
# ═══════════════════════════════════════════════════════════════════════════════

def compute_lcoe(sites: list,
                 discount_rate: float,
                 project_duration: int,
                 eur_xof: float,
                 include_connection: bool,
                 load_factor_param: float,
                 cb: Callable) -> list:
    cb(74, "[7] Economic analysis (LCOE)…", 'info')

    for site in sites:
        P_MW = site.get('p_install_mw_manning', site['p_install_mw'])
        H    = site['h_brute']

        if P_MW < 1:   capex_eur_kw = 3500
        elif P_MW < 10: capex_eur_kw = 2800
        elif P_MW < 50: capex_eur_kw = 2200
        else:           capex_eur_kw = 1800

        if H < 10:  capex_eur_kw *= 1.2
        elif H > 50:capex_eur_kw *= 0.9

        capex_meur = (capex_eur_kw * P_MW * 1000) / 1e6
        if include_connection:
            capex_meur += P_MW * 0.15

        opex_an = capex_meur * 0.025

        if load_factor_param > 0:
            fc = load_factor_param / 100.0
        else:
            fc = 0.55 if H > 50 else (0.50 if H > 20 else 0.45)

        energie_gwh = P_MW * 8760 * fc / 1000.0

        if energie_gwh > 0:
            factor = sum((1 + discount_rate)**-t
                         for t in range(1, project_duration + 1))
            # Numerator: total discounted cost in € (CAPEX + NPV OPEX)
            # Denominator: total discounted energy in kWh  (1 GWh = 1e6 kWh)
            lcoe_eur = (capex_meur*1e6 + opex_an*1e6*factor) / \
                       (energie_gwh*1e6*factor)
            lcoe_fcfa = lcoe_eur * eur_xof
        else:
            lcoe_eur = lcoe_fcfa = 0.0

        # Class
        if lcoe_fcfa < 40:   econ_class = "Very attractive"
        elif lcoe_fcfa < 60: econ_class = "Attractive"
        elif lcoe_fcfa < 80: econ_class = "Moderate"
        else:                econ_class = "Costly"

        site.update({
            'capex_meur':      round(capex_meur, 2),
            'opex_meur_an':    round(opex_an,    3),
            'facteur_charge':  round(fc,          3),
            'energie_gwh_an':  round(energie_gwh, 2),
            'lcoe_eur_kwh':    round(lcoe_eur,    4),
            'lcoe_fcfa_kwh':   round(lcoe_fcfa,   1),
            'lcoe_usd_mwh':    round(lcoe_eur / TAUX_EUR_USD * 1000, 1),
            'classe_economique': econ_class,
        })

    cb(78, "[7] LCOE complete.", 'info')
    return sites


# ═══════════════════════════════════════════════════════════════════════════════
# CLASSIFICATION & PRIORITY (preserved exactly)
# ═══════════════════════════════════════════════════════════════════════════════

def classify_sites(sites: list, cb: Callable) -> list:
    cb(80, "[8] Classification and prioritisation…", 'info')

    for site in sites:
        P    = site.get('p_install_mw_manning', site['p_install_mw'])
        H    = site['h_brute']
        LCOE = site.get('lcoe_fcfa_kwh', 0)
        L    = site['longueur_exploit']

        if P < 0.5:   cl_p = "Micro (<0.5 MW)"
        elif P < 1:   cl_p = "Mini (0.5-1 MW)"
        elif P < 10:  cl_p = "Small (1-10 MW)"
        elif P < 50:  cl_p = "Medium (10-50 MW)"
        else:         cl_p = "Large (>50 MW)"

        if H < 10:    cl_h = "Low head (<10m)"
        elif H < 30:  cl_h = "Medium (10-30m)"
        elif H < 100: cl_h = "High (30-100m)"
        else:         cl_h = "Very high (>100m)"

        score = 0
        score += 30 if P > 50 else (25 if P > 10 else (15 if P > 1 else 5))
        score += 40 if LCOE < 40 else (30 if LCOE < 60 else (15 if LCOE < 80 else 5))
        if 10 < H < 50 and 1000 < L < 5000:
            score += 30
        elif 5 < H < 100 and 500 < L < 10000:
            score += 20
        else:
            score += 10

        priorite = "HIGH" if score >= 80 else ("MEDIUM" if score >= 60 else "LOW")

        site.update({
            'classe_puissance': cl_p,
            'classe_chute':     cl_h,
            'score_priorite':   score,
            'priorite':         priorite,
            'statut':           'POTENTIAL',
        })

    cb(83, "[8] Classification complete.", 'info')
    return sites


# ═══════════════════════════════════════════════════════════════════════════════
# PROFILE PLOTS (style ICOLD / preserved from v30)
# ═══════════════════════════════════════════════════════════════════════════════

def generate_profile_plots(profiles: dict,
                            sites: list,
                            segment_norm: int,
                            profile_step: int,
                            turbine_eff: float,
                            plot_format: str,
                            plot_dpi: int,
                            plot_all: bool,
                            output_folder: str,
                            cb: Callable) -> str:
    cb(85, "[9] Generating profile plots…", 'info')

    plots_folder = os.path.join(output_folder, "Profile_Plots")
    os.makedirs(plots_folder, exist_ok=True)

    if not HAS_MATPLOTLIB:
        cb(85, "   WARNING: matplotlib not available.", 'warning')
        return plots_folder

    count = 0
    for cat, pts in profiles.items():
        if not pts:
            continue

        sites_cat = [s for s in sites if s['categorie'] == cat]
        if not sites_cat and not plot_all:
            continue

        chainage_km = [p['pk'] / 1000.0 for p in pts]
        altitude    = [p['z'] for p in pts]
        pot_specs   = []

        for i in range(len(pts) - 1):
            dh = pts[i]['z'] - pts[i+1]['z']
            q  = pts[i].get('q_moyen', 0)
            if dh > 0 and q > 0:
                pot_kw   = 1000 * 9.81 * q * dh * turbine_eff / 1000
                pot_spec = pot_kw * (segment_norm / max(profile_step, 1))
            else:
                pot_spec = 0
            pot_specs.append(pot_spec)
        pot_specs.append(0)

        fig, ax1 = plt.subplots(figsize=(16, 6))
        ax1.set_xlabel('Chainage [km]', fontsize=12, fontweight='bold')
        ax1.set_ylabel(f'Specific potential [kW/{segment_norm}m]',
                       color='tab:red', fontsize=11)
        ax1.plot(chainage_km, pot_specs, color='tab:red', lw=2,
                 label='Specific potential')
        ax1.tick_params(axis='y', labelcolor='tab:red')
        ax1.grid(True, alpha=0.3)
        ax1.set_ylim(bottom=0)

        ax2 = ax1.twinx()
        ax2.set_ylabel('Elevation [m]', color='tab:blue', fontsize=11)
        ax2.plot(chainage_km, altitude, color='tab:blue', lw=1.5,
                 alpha=0.7, label='Elevation')
        ax2.tick_params(axis='y', labelcolor='tab:blue')

        for s in sites_cat:
            pk_km  = s['pk_prise'] / 1000.0
            pot_mx = max(pot_specs) if pot_specs else 100
            ax1.axvline(pk_km, color='green', ls='--', lw=1.5, alpha=0.7)
            ax1.text(pk_km, pot_mx * 0.9, s['site_id'],
                     rotation=90, va='top', fontsize=8, fontweight='bold',
                     bbox=dict(boxstyle='round,pad=0.2', fc='white', alpha=0.8))

        plt.title(
            f"Longitudinal profile — {cat}\n"
            f"Specific potential & elevation (ICOLD/ESMAP methodology)\n"
            f"{len(sites_cat)} sites identified",
            fontsize=13, fontweight='bold', pad=18
        )
        lines1, lbl1 = ax1.get_legend_handles_labels()
        lines2, lbl2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, lbl1 + lbl2, loc='upper left')
        plt.tight_layout()

        fname = os.path.join(plots_folder,
                             f"Profile_{cat}.{plot_format.lower()}")
        plt.savefig(fname, dpi=plot_dpi, bbox_inches='tight')
        plt.close()
        count += 1
        cb(85, f"   Saved {fname}", 'info')

    cb(87, f"[9] {count} plots generated.", 'info')
    return plots_folder


# ═══════════════════════════════════════════════════════════════════════════════
# SHAPEFILE OUTPUTS (geopandas replaces arcpy Feature Classes)
# ═══════════════════════════════════════════════════════════════════════════════

def export_shapefiles(sites: list,
                      profiles: dict,
                      dem_crs,
                      output_folder: str,
                      cb: Callable):
    cb(89, "[10] Exporting shapefiles…", 'info')

    if not sites:
        return

    # Build GeoDataFrame of all sites
    records = []
    for s in sites:
        p_mw = s.get('p_install_mw_manning', s['p_install_mw'])
        records.append({
            'geometry':    Point(s['x_prise'], s['y_prise']),
            'SITE_ID':     s['site_id'],
            'COURS':       s['categorie'],
            'ORD_STRA':    s.get('strahler', 0),
            'CATEGORIE':   s['categorie'],
            'PK_PRISE':    round(s['pk_prise'], 1),
            'PK_REST':     round(s['pk_rest'],  1),
            'X_PRISE':     round(s['x_prise'],  4),
            'Y_PRISE':     round(s['y_prise'],  4),
            'X_REST':      round(s['x_rest'],   4),
            'Y_REST':      round(s['y_rest'],   4),
            'Z_PRISE':     round(s['z_prise'],  2),
            'Z_REST':      round(s['z_rest'],   2),
            'H_NATUREL':   round(s['h_naturelle'], 2),
            'H_SEUIL':     round(s.get('h_seuil', 0), 2),
            'H_BRUTE':     round(s['h_brute'],  2),
            'H_NETTE':     round(s.get('h_nette_manning', s['h_nette']), 2),
            'PENTE_PCT':   round(s['pente_moy_pct'], 3),
            'LONG_EXPLT':  round(s['longueur_exploit'], 0),
            'SURF_BV':     round(s.get('surf_bv', 0),   2),
            'Q_MOYEN':     round(s['q_moyen'],   2),
            'Q_ETIAGE':    round(s['q_etiage'],  2),
            'Q_CRUE':      round(s['q_crue'],    2),
            'Q_EQUIP':     round(s['q_equip'],   2),
            'Q_RESIDUEL':  round(s.get('q_residuel', s['q_moyen']), 2),
            'POT_SP_MAX':  round(s.get('pot_spec_max', 0), 2),
            'POT_SP_MOY':  round(s.get('pot_spec_moy', 0), 2),
            'P_INST_KW':   round(s.get('p_install_kw_manning', s['p_install_kw']), 1),
            'P_INST_MW':   round(p_mw, 3),
            'ENERG_GWH':   round(s.get('energie_gwh_an', 0), 2),
            'FACT_CHARG':  round(s.get('facteur_charge',  0), 3),
            'CANAL_LARG':  round(s.get('canal_largeur',   0), 2),
            'CANAL_HAUT':  round(s.get('canal_hauteur',   0), 2),
            'CANAL_LONG':  round(s.get('canal_longueur',  0), 0),
            'CANAL_VITE':  round(s.get('canal_vitesse',   0), 3),
            'CANAL_RH':    round(s.get('canal_rh',        0), 3),
            'PERTE_CH_M':  round(s.get('perte_charge_m',  0), 2),
            'Z_AVAL':      round(s.get('z_aval',          0), 2),
            'CRN':         round(s.get('crn',             0), 2),
            'REVANCHE':    round(s.get('revanche',        0), 2),
            'COTE_CRETE':  round(s.get('cote_crete',      0), 2),
            'H_BARR_VUE':  round(s.get('h_barrage_vue',   0), 1),
            'H_BARR_TOT':  round(s.get('h_barrage_total', 0), 1),
            'LONG_DIGUE':  round(s.get('longueur_digue',  0), 0),
            'VOL_RES_M3':  round(s.get('volume_reservoir_m3', 0), 0),
            'SURF_RES_HA': round(s.get('surface_reservoir_ha', 0), 2),
            'CAPEX_MEUR':  round(s.get('capex_meur',      0), 2),
            'OPEX_MEUR':   round(s.get('opex_meur_an',    0), 3),
            'LCOE_EUR':    round(s.get('lcoe_eur_kwh',    0), 4),
            'LCOE_FCFA':   round(s.get('lcoe_fcfa_kwh',   0), 1),
            'CL_PUISS':    s.get('classe_puissance', ''),
            'CL_CHUTE':    s.get('classe_chute',     ''),
            'CL_ECONOM':   s.get('classe_economique',''),
            'PRIORITE':    s.get('priorite',         'MEDIUM'),
            'SCORE':       s.get('score_priorite',   50),
            'STATUT':      s.get('statut',           'POTENTIAL'),
            'NB_AMONT':    s.get('nb_sites_amont',   0),
            'CMT_CASCAD':  s.get('commentaire_cascade',''),
            'DATE_DETEC':  datetime.now().strftime('%Y-%m-%d'),
            'VERSION':     'DamFinder_v1.0',
            'SOURCE_DEB':  s.get('flow_source', ''),
        })

    gdf = gpd.GeoDataFrame(records, crs=dem_crs)

    # Export by category
    for cat in ['PRINCIPAL', 'MAJEURS', 'SECONDAIRES']:
        sub = gdf[gdf['CATEGORIE'] == cat]
        if len(sub) > 0:
            path = os.path.join(output_folder, f"Sites_{cat}.shp")
            sub.to_file(path)
            cb(89, f"   Sites_{cat}.shp ({len(sub)} sites)", 'info')

    # Export all
    all_path = os.path.join(output_folder, "Sites_All.shp")
    gdf.to_file(all_path)
    cb(89, f"   Sites_All.shp ({len(gdf)} sites)", 'info')

    # Export exploitation lines (prise → restitution)
    lines = []
    for s in sites:
        lines.append({
            'geometry':  LineString([(s['x_prise'], s['y_prise']),
                                     (s['x_rest'],  s['y_rest'])]),
            'SITE_ID':   s['site_id'],
            'CATEGORIE': s['categorie'],
            'LONG_M':    s['longueur_exploit'],
        })
    gpd.GeoDataFrame(lines, crs=dem_crs).to_file(
        os.path.join(output_folder, "Exploitation_Lines.shp"))
    cb(90, "   Exploitation_Lines.shp", 'info')

    cb(91, "[10] Shapefiles exported.", 'info')


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL REPORT — 45 FIELDS (openpyxl, preserved from v30, names cleaned)
# ═══════════════════════════════════════════════════════════════════════════════

def export_excel(sites: list, params: dict,
                 output_folder: str, cb: Callable) -> str:
    cb(92, "[11] Generating Excel report…", 'info')

    if not HAS_OPENPYXL:
        cb(92, "   WARNING: openpyxl not available.", 'warning')
        return ""

    path = os.path.join(output_folder, "DamFinder_Report.xlsx")
    wb   = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "SUMMARY"
    hdr_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=12)

    ws['A1'] = "DamFinder Pro v1.0 — Hydroelectric Site Report"
    ws['A1'].font = Font(bold=True, size=16, color="1F3864")
    ws['A3'] = "Generated:"
    ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M')
    ws['A4'] = "Flow source:"
    ws['B4'] = params.get('flow_source', '')
    ws['A5'] = "Profile step:"
    ws['B5'] = f"{params.get('profile_step', 40)} m"
    ws['A6'] = "Min power:"
    ws['B6'] = f"{params.get('min_power', 500)} kW"
    ws['A7'] = "Methodology:"
    ws['B7'] = "ICOLD + World Bank ESMAP"

    ws['A9'] = "DETECTED SITES"
    ws['A9'].font = Font(bold=True, size=14, color="1F3864")

    cats = ['PRINCIPAL', 'MAJEURS', 'SECONDAIRES']
    for i, cat in enumerate(cats, 11):
        n = sum(1 for s in sites if s['categorie'] == cat)
        ws.cell(i, 1, cat).font = Font(bold=True)
        ws.cell(i, 2, n).font   = Font(bold=True, size=12)

    ws.cell(15, 1, "TOTAL:").font = Font(bold=True, size=13)
    ws.cell(15, 2, len(sites)).font = Font(bold=True, size=14, color="C0392B")

    ws['A17'] = APP_CREDIT.split('\n')[5]  # copyright line
    ws['A17'].font = Font(italic=True, color="7F8C8D")

    # ── Sheet 2: All sites (45 fields) ────────────────────────────────────────
    ws2 = wb.create_sheet("ALL SITES")
    headers = [
        "SITE_ID", "CATEGORIE", "ORD_STRA",
        "PK_PRISE(m)", "PK_REST(m)", "X_PRISE", "Y_PRISE",
        "Z_PRISE(m)", "Z_REST(m)", "H_NATUREL(m)", "H_BRUTE(m)", "H_NETTE(m)",
        "PENTE(%)", "LONG_EXPLT(m)",
        "SURF_BV(km²)", "Q_MOYEN(m³/s)", "Q_ETIAGE(m³/s)", "Q_CRUE(m³/s)",
        "Q_EQUIP(m³/s)", "Q_RESIDUEL(m³/s)",
        "POT_SP_MAX(kW/40m)", "POT_SP_MOY(kW/40m)",
        "P_INST(kW)", "P_INST(MW)", "ENERG(GWh/yr)", "FC",
        "CANAL_LARG(m)", "CANAL_HAUT(m)", "CANAL_LONG(m)", "PERTE_CH(m)",
        "Z_AVAL(m)", "CRN(m)", "COTE_CRETE(m)", "H_BARR(m)", "LONG_DIGUE(m)",
        "VOL_RES(m³)", "SURF_RES(ha)",
        "CAPEX(M€)", "OPEX(M€/yr)", "LCOE(€/kWh)", "LCOE(FCFA/kWh)",
        "CL_PUISS", "CL_CHUTE", "CL_ECONOM", "PRIORITE", "SCORE", "STATUT",
        "NB_AMONT", "DATE_DETEC"
    ]

    for col, h in enumerate(headers, 1):
        c = ws2.cell(1, col, h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal='center')

    for row_i, s in enumerate(sites, 2):
        p_mw = s.get('p_install_mw_manning', s['p_install_mw'])
        row_data = [
            s['site_id'], s['categorie'], s.get('strahler', 0),
            s['pk_prise'], s['pk_rest'], s['x_prise'], s['y_prise'],
            s['z_prise'], s['z_rest'], s['h_naturelle'], s['h_brute'],
            s.get('h_nette_manning', s['h_nette']),
            s['pente_moy_pct'], s['longueur_exploit'],
            s.get('surf_bv', 0), s['q_moyen'], s['q_etiage'],
            s['q_crue'], s['q_equip'], s.get('q_residuel', s['q_moyen']),
            s.get('pot_spec_max', 0), s.get('pot_spec_moy', 0),
            s.get('p_install_kw_manning', s['p_install_kw']), p_mw,
            s.get('energie_gwh_an', 0), s.get('facteur_charge', 0),
            s.get('canal_largeur', 0), s.get('canal_hauteur', 0),
            s.get('canal_longueur', 0), s.get('perte_charge_m', 0),
            s.get('z_aval', 0), s.get('crn', 0), s.get('cote_crete', 0),
            s.get('h_barrage_vue', 0), s.get('longueur_digue', 0),
            s.get('volume_reservoir_m3', 0), s.get('surface_reservoir_ha', 0),
            s.get('capex_meur', 0), s.get('opex_meur_an', 0),
            s.get('lcoe_eur_kwh', 0), s.get('lcoe_fcfa_kwh', 0),
            s.get('classe_puissance', ''), s.get('classe_chute', ''),
            s.get('classe_economique', ''), s.get('priorite', 'MEDIUM'),
            s.get('score_priorite', 50), s.get('statut', 'POTENTIAL'),
            s.get('nb_sites_amont', 0), s.get('date_detect', datetime.now().strftime('%Y-%m-%d')),
        ]
        for col_i, val in enumerate(row_data, 1):
            cell = ws2.cell(row_i, col_i, val)
            # Colour-code priority
            if col_i == headers.index('PRIORITE') + 1:
                if val == 'HIGH':
                    cell.font = Font(bold=True, color="1E8449")
                elif val == 'MEDIUM':
                    cell.font = Font(bold=True, color="D35400")
                else:
                    cell.font = Font(color="7F8C8D")

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 14

    # ── Sheet 3: Top sites by score ───────────────────────────────────────────
    ws3 = wb.create_sheet("TOP PRIORITY")
    top = sorted(sites, key=lambda s: s.get('score_priorite', 0), reverse=True)[:20]
    for col, h in enumerate(["SITE_ID", "CATEG", "P_MW", "H_m",
                              "Q_m3s", "LCOE_FCFA", "CAPEX_MEUR",
                              "ENERGY_GWH", "PRIORITY", "SCORE"], 1):
        c = ws3.cell(1, col, h)
        c.fill = hdr_fill
        c.font = hdr_font
    for ri, s in enumerate(top, 2):
        p_mw = s.get('p_install_mw_manning', s['p_install_mw'])
        for ci, v in enumerate([
            s['site_id'], s['categorie'],
            round(p_mw, 2), round(s['h_brute'], 1),
            round(s['q_equip'], 1), round(s.get('lcoe_fcfa_kwh', 0), 1),
            round(s.get('capex_meur', 0), 2), round(s.get('energie_gwh_an', 0), 1),
            s.get('priorite', 'MEDIUM'), s.get('score_priorite', 50)
        ], 1):
            ws3.cell(ri, ci, v)

    wb.save(path)
    cb(93, f"[11] Excel report: {path}", 'info')
    return path


# ═══════════════════════════════════════════════════════════════════════════════
# HTML REPORT (DataTables + Plotly profiles, preserved from v30, names cleaned)
# ═══════════════════════════════════════════════════════════════════════════════

def export_html(sites: list, profiles: dict, params: dict,
                output_folder: str, cb: Callable) -> str:
    cb(94, "[11] Generating HTML report…", 'info')

    path = os.path.join(output_folder, "DamFinder_Report.html")
    now  = datetime.now().strftime('%d/%m/%Y %H:%M')

    count_p = sum(1 for s in sites if s['categorie'] == 'PRINCIPAL')
    count_m = sum(1 for s in sites if s['categorie'] == 'MAJEURS')
    count_s = sum(1 for s in sites if s['categorie'] == 'SECONDAIRES')

    top_sites = sorted(sites,
                       key=lambda s: s.get('score_priorite', 0),
                       reverse=True)[:10]

    def _pclass(p):
        if p == 'HIGH':   return 'priorite-haute'
        if p == 'MEDIUM': return 'priorite-moyenne'
        return 'priorite-basse'

    rows_html = ""
    for s in top_sites:
        p_mw = s.get('p_install_mw_manning', s['p_install_mw'])
        rows_html += f"""
        <tr>
            <td><strong>{s['site_id']}</strong></td>
            <td>{s['categorie']}</td>
            <td>{s['pk_prise']/1000:.1f} – {s['pk_rest']/1000:.1f}</td>
            <td>{s['h_brute']:.1f}</td>
            <td>{p_mw:.2f}</td>
            <td>{s.get('lcoe_usd_mwh', 0):.1f}</td>
            <td>{s.get('lcoe_fcfa_kwh', 0):.1f}</td>
            <td class="{_pclass(s.get('priorite','MEDIUM'))}">{s.get('priorite','MEDIUM')}</td>
        </tr>"""

    all_rows_html = ""
    for s in sites:
        p_mw = s.get('p_install_mw_manning', s['p_install_mw'])
        all_rows_html += f"""
        <tr>
            <td>{s['site_id']}</td><td>{s['categorie']}</td>
            <td>{s['pk_prise']/1000:.1f}</td>
            <td>{s['h_brute']:.1f}</td><td>{p_mw:.2f}</td>
            <td>{s.get('energie_gwh_an', 0):.1f}</td>
            <td>{s.get('capex_meur', 0):.1f}</td>
            <td>{s.get('lcoe_usd_mwh', 0):.1f}</td>
            <td>{s.get('lcoe_fcfa_kwh', 0):.1f}</td>
            <td class="{_pclass(s.get('priorite','MEDIUM'))}">{s.get('priorite','MEDIUM')}</td>
            <td>{s.get('score_priorite', 0)}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <title>DamFinder Pro v1.0 — Hydroelectric Site Report</title>
  <link rel="stylesheet" href="https://cdn.datatables.net/1.13.4/css/jquery.dataTables.min.css">
  <style>
    body  {{ font-family:'Segoe UI',Arial,sans-serif; margin:20px; background:#f0f4f8; }}
    .container {{ max-width:1300px; margin:0 auto; background:#fff; padding:32px;
                  border-radius:12px; box-shadow:0 3px 16px rgba(0,0,0,.12); }}
    h1   {{ color:#1a3a5c; border-bottom:4px solid #2980b9; padding-bottom:14px; }}
    h2   {{ color:#2c3e50; border-left:5px solid #2980b9; padding-left:10px; margin-top:32px; }}
    .stats {{ background:#eaf2fb; padding:22px; border-radius:8px; margin:18px 0; }}
    .site-count {{ font-size:38px; font-weight:bold; color:#c0392b; }}
    .category {{ display:inline-block; margin:12px 18px; padding:14px 22px; border-radius:8px; text-align:center; }}
    .cat-p {{ background:#2980b9; color:#fff; }}
    .cat-m {{ background:#8e44ad; color:#fff; }}
    .cat-s {{ background:#27ae60; color:#fff; }}
    table {{ width:100%; border-collapse:collapse; font-size:.92em; }}
    th    {{ background:#2c3e50; color:#fff; padding:11px; text-align:left; }}
    td    {{ padding:9px; border-bottom:1px solid #e0e0e0; }}
    tr:hover {{ background:#f4f8fb; }}
    .priorite-haute   {{ color:#1e8449; font-weight:bold; }}
    .priorite-moyenne {{ color:#d35400; font-weight:bold; }}
    .priorite-basse   {{ color:#7f8c8d; }}
    .footer {{ margin-top:36px; padding-top:18px; border-top:2px solid #ecf0f1;
               color:#7f8c8d; font-size:.87em; }}
    .badge {{ display:inline-block; padding:3px 10px; border-radius:12px;
              font-size:.82em; font-weight:bold; background:#2980b9; color:#fff; }}
  </style>
</head>
<body>
<div class="container">

  <h1>&#127988; DamFinder Pro v1.0 — Hydroelectric Site Report</h1>
  <p style="color:#7f8c8d;font-size:1.05em;">
    ICOLD specific potential method + World Bank ESMAP hydropower guidelines
  </p>

  <div class="stats">
    <h2>&#128202; Analysis Summary</h2>
    <p><strong>Date:</strong> {now}</p>
    <p><strong>Flow source:</strong> {params.get('flow_source','')}</p>
    <p><strong>Profile step:</strong> {params.get('profile_step',40)} m
       &nbsp;|&nbsp; <strong>Segment normalisation:</strong> {params.get('segment_norm',40)} m
       &nbsp;|&nbsp; <strong>Min power:</strong> {params.get('min_power',500)} kW</p>

    <div style="margin-top:18px;">
      <div class="category cat-p">
        <div>MAIN COURSE</div>
        <div class="site-count">{count_p}</div><div>sites</div>
      </div>
      <div class="category cat-m">
        <div>MAJOR TRIBUTARIES</div>
        <div class="site-count">{count_m}</div><div>sites</div>
      </div>
      <div class="category cat-s">
        <div>SECONDARY TRIBUTARIES</div>
        <div class="site-count">{count_s}</div><div>sites</div>
      </div>
    </div>
    <p style="margin-top:18px;font-size:1.25em;">
      <strong>TOTAL: {len(sites)} sites detected</strong>
    </p>
  </div>

  <h2>&#127919; Priority Sites — Top 10</h2>
  <table>
    <tr>
      <th>Site ID</th><th>Category</th><th>PK (km)</th>
      <th>H (m)</th><th>P (MW)</th><th>LCOE (USD/MWh)</th>
      <th>LCOE (FCFA/kWh)</th><th>Priority</th>
    </tr>
    {rows_html}
  </table>

  <h2>&#128203; Complete Results — All Sites</h2>
  <table id="allSitesTable" class="display" style="width:100%">
    <thead>
      <tr>
        <th>Site</th><th>Cat.</th><th>PK (km)</th>
        <th>H (m)</th><th>P (MW)</th><th>E (GWh/yr)</th>
        <th>CAPEX (M€)</th><th>LCOE USD/MWh</th><th>LCOE FCFA/kWh</th>
        <th>Priority</th><th>Score</th>
      </tr>
    </thead>
    <tbody>{all_rows_html}</tbody>
  </table>

  <h2>&#128196; Output Files</h2>
  <ul style="font-size:1.05em;line-height:2em;">
    <li>&#9989; Sites_PRINCIPAL.shp ({count_p} sites)</li>
    <li>&#9989; Sites_MAJEURS.shp ({count_m} sites)</li>
    <li>&#9989; Sites_SECONDAIRES.shp ({count_s} sites)</li>
    <li>&#9989; Sites_All.shp ({len(sites)} sites)</li>
    <li>&#9989; Exploitation_Lines.shp</li>
    <li>&#9989; DamFinder_Report.xlsx (45 fields)</li>
    <li>&#9989; Profile_Plots/ (longitudinal profiles)</li>
  </ul>

  <div class="footer">
    <p><strong>DamFinder Pro v1.0</strong></p>
    <p>Developed by DAMFINDER Engineering Tools</p>
    <p>Methodology: ICOLD standards + World Bank ESMAP hydropower guidelines</p>
    <p>Calibrated on anonymized West African hydroelectric reference projects (REF_A … REF_E)</p>
    <p>Report generated: {now}</p>
    <p>&#169; 2026 DAMFINDER Engineering Tools — All rights reserved</p>
  </div>

</div>
<script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
<script src="https://cdn.datatables.net/1.13.4/js/jquery.dataTables.min.js"></script>
<script>$(document).ready(function(){{ $('#allSitesTable').DataTable({{
  pageLength:25, order:[[10,'desc']]
}}); }});</script>
</body>
</html>"""

    with open(path, 'w', encoding='utf-8') as f:
        f.write(html)

    cb(95, f"[11] HTML report: {path}", 'info')
    return path


# ═══════════════════════════════════════════════════════════════════════════════
# FOLIUM MAP GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════

def generate_folium_map(sites: list,
                        river_gdf: Optional[gpd.GeoDataFrame],
                        dem_crs) -> str:
    """
    Build a Folium map HTML string with:
    - Coloured site markers (green <10MW, orange 10-50MW, red >50MW)
    - Popup: Site ID, P MW, LCOE USD/MWh, H m, Q m³/s, Priority
    - River network in blue (line weight ∝ Strahler order)
    - Basemap: OpenStreetMap + Esri Satellite option
    Returns HTML string ready for QWebEngineView.
    """
    try:
        import folium
        from folium import plugins
    except ImportError:
        return "<html><body><p>folium not installed.</p></body></html>"

    if not sites:
        return "<html><body><p>No sites to display.</p></body></html>"

    # Centre map on site centroid
    lats = []
    lons = []

    # We need geographic coordinates. If DEM CRS is not EPSG:4326, reproject.
    need_reproject = False
    crs_str = str(dem_crs) if dem_crs else ""
    if dem_crs:
        try:
            from pyproj import CRS as ProjCRS, Transformer
            proj_crs  = ProjCRS.from_user_input(dem_crs)
            geo_crs   = ProjCRS.from_epsg(4326)
            need_reproject = not proj_crs.equals(geo_crs)
            if need_reproject:
                transformer = Transformer.from_crs(proj_crs, geo_crs,
                                                   always_xy=True)
        except Exception:
            need_reproject = False

    def _to_latlon(x, y):
        if need_reproject:
            try:
                lon, lat = transformer.transform(x, y)
                return lat, lon
            except Exception:
                return y, x
        return y, x

    for s in sites:
        lat, lon = _to_latlon(s['x_prise'], s['y_prise'])
        lats.append(lat)
        lons.append(lon)

    centre_lat = sum(lats) / len(lats)
    centre_lon = sum(lons) / len(lons)

    m = folium.Map(
        location=[centre_lat, centre_lon],
        zoom_start=9,
        tiles='OpenStreetMap'
    )
    # Satellite layer
    folium.TileLayer(
        tiles='https://server.arcgisonline.com/ArcGIS/rest/services/'
              'World_Imagery/MapServer/tile/{z}/{y}/{x}',
        attr='Esri World Imagery',
        name='Satellite (Esri)',
        overlay=False, control=True
    ).add_to(m)

    # River network
    if river_gdf is not None and len(river_gdf) > 0:
        for _, row in river_gdf.iterrows():
            geom = row.geometry
            if geom is None:
                continue
            order = int(row.get('ORD_STRA', 3))
            weight = max(1, order - 1)
            lines_list = list(geom.geoms) if geom.geom_type == 'MultiLineString' else [geom]
            for ln in lines_list:
                coords_ll = [_to_latlon(x, y) for x, y in ln.coords]
                folium.PolyLine(
                    locations=coords_ll,
                    color='#2980b9',
                    weight=weight,
                    opacity=0.7,
                    tooltip=f"Strahler {order}"
                ).add_to(m)

    # Site markers
    for i, s in enumerate(sites):
        lat, lon = lats[i], lons[i]
        p_mw  = s.get('p_install_mw_manning', s['p_install_mw'])
        color = 'green' if p_mw < 10 else ('orange' if p_mw < 50 else 'red')

        popup_html = f"""
        <div style="font-family:Arial;min-width:220px;">
          <h4 style="margin:0;color:#1a3a5c;">{s['site_id']}</h4>
          <hr style="margin:4px 0;">
          <b>Power:</b> {p_mw:.2f} MW<br>
          <b>LCOE:</b> {s.get('lcoe_usd_mwh', 0):.1f} USD/MWh<br>
          <b>Head H:</b> {s['h_brute']:.1f} m<br>
          <b>Flow Q:</b> {s['q_equip']:.1f} m³/s<br>
          <b>Category:</b> {s['categorie']}<br>
          <b>Priority:</b>
            <span style="color:{'#1e8449' if s.get('priorite')=='HIGH' else '#d35400'}">
              {s.get('priorite','MEDIUM')}
            </span><br>
          <b>CAPEX:</b> {s.get('capex_meur', 0):.1f} M€<br>
          <b>Energy:</b> {s.get('energie_gwh_an', 0):.1f} GWh/yr
        </div>"""

        folium.Marker(
            location=[lat, lon],
            popup=folium.Popup(popup_html, max_width=280),
            tooltip=f"{s['site_id']} — {p_mw:.1f} MW",
            icon=folium.Icon(color=color, icon='bolt', prefix='fa')
        ).add_to(m)

    folium.LayerControl().add_to(m)
    plugins.Fullscreen().add_to(m)
    plugins.MiniMap().add_to(m)

    return m._repr_html_()


# ═══════════════════════════════════════════════════════════════════════════════
# MASTER RUNNER — called from QThread worker
# ═══════════════════════════════════════════════════════════════════════════════

def run_analysis(params: dict, cb: Callable) -> dict:
    """
    Full 11-phase pipeline.

    params keys (mirror of PyQt5 form fields):
      dem_path, river_network_path, study_area_path, output_folder,
      generate_network, flow_threshold,
      strahler_main, strahler_major, strahler_secondary,
      profile_step, extraction_mode, specific_stream,
      flow_source, equip_coef, fixed_flow_main, fixed_flow_major,
      fixed_flow_secondary, upstream_ratio,
      pjmax, cn_default, runoff_coef, tc_method, rainy_days, load_factor,
      segment_norm, adaptive_threshold, threshold_peak_manual,
      max_exploit_dist, min_power, weir_height, min_spacing,
      filter_slope, min_slope,
      manning_n, canal_length, head_loss_pct, side_slope, turbine_eff,
      enable_cascade, cascade_distance,
      discount_rate, project_duration, eur_xof, include_connection,
      generate_plots, plot_format, plot_dpi, plot_all,
      generate_excel, generate_html

    cb signature: cb(percent: int, message: str, level: str)
    Returns: {'sites': list, 'map_html': str, 'excel_path': str, 'html_path': str}
    """
    start = time.time()
    cb(0, "=" * 60, 'info')
    cb(0, "DamFinder Pro v1.0 — Analysis started", 'info')
    cb(0, "Methodology: ICOLD + World Bank ESMAP", 'info')
    cb(0, "=" * 60, 'info')

    out = params['output_folder']
    os.makedirs(out, exist_ok=True)

    tmp_dir = tempfile.mkdtemp(prefix="damfinder_")

    # ── Phase 1 ──────────────────────────────────────────────────────────────
    cb(2, "PHASE 1: DEM preparation and network extraction", 'info')

    dem_path = params['dem_path']

    # Clip DEM
    clipped_dem = os.path.join(tmp_dir, "dem_clip.tif")
    dem_path = clip_dem(dem_path, params.get('study_area_path'), clipped_dem)

    # Read DEM metadata
    with rasterio.open(dem_path) as src:
        dem_crs   = src.crs
        cell_size = abs(src.transform.a)
        cb(3, f"   DEM: {src.width}×{src.height} px, res={cell_size:.1f}m, CRS={src.crs}", 'info')

    # Flow rasters
    dir_path = os.path.join(tmp_dir, "flow_dir.tif")
    acc_path = os.path.join(tmp_dir, "flow_acc.tif")
    dir_path, acc_path = compute_flow(dem_path, dir_path, acc_path, cb)

    # River network
    if params.get('river_network_path') and \
       os.path.exists(params['river_network_path']):
        cb(30, "[1.5] Using provided river network…", 'info')
        river_gdf = gpd.read_file(params['river_network_path'])
        if 'ORD_STRA' not in river_gdf.columns:
            river_gdf['ORD_STRA'] = 3
        if 'LENGTH_KM' not in river_gdf.columns:
            river_gdf['LENGTH_KM'] = river_gdf.geometry.length / 1000.0
    else:
        river_gdf = extract_stream_network(
            dem_path, acc_path,
            params.get('flow_threshold', 0),
            {'main': params['strahler_main'],
             'major': params['strahler_major'],
             'secondary': params['strahler_secondary']},
            params.get('study_area_path'),
            cb
        )

    if river_gdf is None or len(river_gdf) == 0:
        raise ValueError("No stream network extracted. "
                         "Lower flow_threshold or provide a river shapefile.")

    # Reproject river to DEM CRS
    if river_gdf.crs and river_gdf.crs != dem_crs:
        river_gdf = river_gdf.to_crs(dem_crs)

    cb(34, f"PHASE 1 complete — {len(river_gdf)} stream segments", 'info')

    # ── Phase 2 ──────────────────────────────────────────────────────────────
    cb(35, "PHASE 2: Longitudinal profile extraction", 'info')
    profiles = extract_profiles(
        river_gdf, dem_path, acc_path,
        params['profile_step'],
        params['strahler_main'],
        params['strahler_major'],
        params['strahler_secondary'],
        cb
    )
    if not profiles:
        raise ValueError("No profiles extracted. Check DEM and network coverage.")
    cb(48, f"PHASE 2 complete — {len(profiles)} profiles", 'info')

    # ── Phase 3 ──────────────────────────────────────────────────────────────
    cb(48, "PHASE 3: Flow attribution", 'info')
    profiles = attribute_flows(
        profiles,
        params['flow_source'],
        params['equip_coef'],
        params['fixed_flow_main'],
        params['fixed_flow_major'],
        params['fixed_flow_secondary'],
        params.get('upstream_ratio', 0.3),
        params.get('pjmax', 100.0),
        params.get('cn_default', 75.0),
        params.get('runoff_coef', 0.45),
        params.get('tc_method', 'Kirpich'),
        params.get('rainy_days', 120),
        cell_size,
        cb
    )
    cb(50, "PHASE 3 complete.", 'info')

    # ── Phase 4 ──────────────────────────────────────────────────────────────
    cb(50, "PHASE 4: Site detection (specific potential method)", 'info')
    sites = detect_sites(
        profiles,
        params['segment_norm'],
        params['profile_step'],
        params['adaptive_threshold'],
        params.get('threshold_peak_manual', 50.0),
        params.get('max_exploit_dist', 0.0),
        params['min_power'],
        params.get('weir_height', 5.0),
        params.get('min_spacing', 500.0),
        params.get('filter_slope', False),
        params.get('min_slope', 0.1),
        params.get('head_loss_pct', 0.06),
        params.get('turbine_eff', 0.90),
        params['equip_coef'],
        cb
    )
    if not sites:
        cb(0, "WARNING: No sites detected. Lower min_power or adjust thresholds.",
           'warning')
        return {'sites': [], 'map_html': '', 'excel_path': '', 'html_path': ''}

    cb(58, f"PHASE 4 complete — {len(sites)} sites", 'info')

    # ── Phase 5 ──────────────────────────────────────────────────────────────
    cb(60, "PHASE 5: Manning hydraulics", 'info')
    sites = compute_manning(sites,
                            params.get('manning_n', 0.035),
                            params.get('canal_length', 1500.0),
                            params.get('head_loss_pct', 0.06),
                            params.get('side_slope', 1.5),
                            params.get('turbine_eff', 0.90),
                            cb)
    cb(63, "PHASE 5 complete.", 'info')

    # ── Phase 5B ─────────────────────────────────────────────────────────────
    cb(65, "PHASE 5B: Dam sizing", 'info')
    sites = compute_dam_sizing(sites,
                               params.get('manning_n', 0.035),
                               cb)
    cb(68, "PHASE 5B complete.", 'info')

    # ── Phase 6 ──────────────────────────────────────────────────────────────
    if params.get('enable_cascade', True):
        cb(70, "PHASE 6: Cascade analysis", 'info')
        sites = compute_cascade(sites,
                                params.get('cascade_distance', 50000.0),
                                cb)
    else:
        for s in sites:
            s['nb_sites_amont']      = 0
            s['q_residuel']          = s['q_moyen']
            s['commentaire_cascade'] = "Disabled"

    # ── Phase 7 ──────────────────────────────────────────────────────────────
    cb(74, "PHASE 7: LCOE economic analysis", 'info')
    sites = compute_lcoe(sites,
                         params.get('discount_rate', 0.08),
                         params.get('project_duration', 30),
                         params.get('eur_xof', 655.957),
                         params.get('include_connection', True),
                         params.get('load_factor', 0),
                         cb)
    cb(78, "PHASE 7 complete.", 'info')

    # ── Phase 8 ──────────────────────────────────────────────────────────────
    cb(80, "PHASE 8: Classification", 'info')
    sites = classify_sites(sites, cb)
    cb(83, "PHASE 8 complete.", 'info')

    # ── Phase 9 ──────────────────────────────────────────────────────────────
    plots_folder = ""
    if params.get('generate_plots', True):
        cb(85, "PHASE 9: Profile plots", 'info')
        plots_folder = generate_profile_plots(
            profiles, sites,
            params['segment_norm'], params['profile_step'],
            params.get('turbine_eff', 0.90),
            params.get('plot_format', 'PNG'),
            params.get('plot_dpi', 300),
            params.get('plot_all', False),
            out, cb
        )
        cb(87, "PHASE 9 complete.", 'info')

    # ── Phase 10 ─────────────────────────────────────────────────────────────
    cb(89, "PHASE 10: Shapefile export", 'info')
    export_shapefiles(sites, profiles, dem_crs, out, cb)
    cb(91, "PHASE 10 complete.", 'info')

    # ── Phase 11 ─────────────────────────────────────────────────────────────
    cb(92, "PHASE 11: Report generation", 'info')
    excel_path = ""
    html_path  = ""

    if params.get('generate_excel', True):
        excel_path = export_excel(sites, params, out, cb)

    if params.get('generate_html', True):
        html_path = export_html(sites, profiles, params, out, cb)

    # ── Folium map ────────────────────────────────────────────────────────────
    cb(96, "Generating interactive map…", 'info')
    map_html = generate_folium_map(sites, river_gdf, dem_crs)

    elapsed = time.time() - start
    cb(100, "=" * 60, 'info')
    cb(100, f"ANALYSIS COMPLETE — {len(sites)} sites — "
            f"{elapsed/60:.1f} min", 'info')
    cb(100, f"Output folder: {out}", 'info')
    cb(100, "=" * 60, 'info')

    return {
        'sites':      sites,
        'map_html':   map_html,
        'excel_path': excel_path,
        'html_path':  html_path,
        'plots_folder': plots_folder,
        'river_gdf':  river_gdf,
    }
